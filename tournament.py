#!/usr/bin/env python3

import json
import csv
import re
import io
import logging
import datetime
import sqlite3
import os.path

import tornado.web
import openpyxl
import tempfile
import handler
import db
import seating
import settings
import util

log = logging.getLogger('WebServer')

def getTournaments(tournamentID=None):
    """Get list of tournament dictionaries. If a tournament ID is provided
    the list will either have 1 dictionary, or be empty if the ID is bad.
    The fields in the dictionary are all the tournament fields in the
    Tournaments table + Code and Flag_Image for the country, a field
    for players that has a dictionary that maps keys of each type of player
    to lists of player dictionaries with Id, Name, Association, and
    NumberScores fields
    """
    tmt_fields = db.table_field_names('Tournaments')
    country_fields = ['Code', 'Flag_Image']
    with db.getCur() as cur:
        sql = """
        SELECT {} FROM Tournaments JOIN Countries ON Country = Countries.ID
        """.format(','.join(['Tournaments.{}'.format(f) for f in tmt_fields] +
                   country_fields))
        if tournamentID:
            sql += "WHERE Tournaments.ID = {}".format(tournamentID)
        sql += " ORDER BY End DESC"
        cur.execute(sql)
        tournaments = [dict(zip(tmt_fields + country_fields, row))
                       for row in cur.fetchall()]
    for tmt in tournaments:
        tmt['Dates'] = '{} - {}'.format(tmt['Start'], tmt['End'])
        tmt['players'] = dict((t, list()) for t in db.playertypes)
        for compete in getCompetitors(tmt['Id']):
            tmt['players'][db.playertypes[compete['Type']]].append(compete)
    return tournaments

def getCompetitors(tournamentID):
    player_fields = ['Name', 'Association']
    cFields = ['Compete.{}'.format(f)
               for f in db.table_field_names('Compete')] + [
                       'Players.{}'.format(f) for f in player_fields]
    sql = """
    SELECT {}, COUNT(DISTINCT Scores.Id)
      FROM Compete JOIN Players on Player = Players.Id
        LEFT OUTER JOIN Rounds ON Rounds.Tournament = Compete.Tournament
        LEFT OUTER JOIN Scores
          ON Scores.Round = Rounds.Id AND Scores.PlayerId = Players.Id
    WHERE Compete.Tournament = ?
    GROUP BY Player
    """.format(','.join(cFields))
    args = (tournamentID,)
    with db.getCur() as cur:
        cur.execute(sql, args)
        competitors = [dict(zip(
            map(db.fieldname, cFields + ['NumberScores']), row))
                       for row in cur.fetchall()]
    return competitors

class TournamentsHandler(handler.BaseHandler):
    def get(self):
        tournaments = getTournaments()
        return self.render("tournamentlist.html", tournaments=tournaments)

class TournamentListHandler(handler.BaseHandler):
    def get(self):
        result = {'status': 0, 'data': getTournaments()}
        return self.write(json.dumps(result))

    @tornado.web.authenticated
    def post(self):
        encoded_item = self.get_argument('item', None)
        item = json.loads(encoded_item)
        result = {'status': 0, 'message': ''}
        tmt_fields = db.table_field_names('Tournaments')
        columns = [f for f in tmt_fields if f in item and f not in ['Id']]
        cleanTournamentItem(item, columns, self.current_user)
        if item.get('Id', None) is None or not isinstance(item['Id'], int):
            result['message'] = 'Invalid Id field for tournament, {}'.format(
                item)
            result['status'] = -1
            return self.write(result)
        if (not self.get_is_admin() and item['Id'] != 0 and 
            str(item.get('Owner', -1)) != self.current_user):
            result['message'] = 'Only owners and admins may edit tournaments'
            result['status'] = -2
            return self.write(result)
        if (not isinstance(item.get('Country', None), int) or
            not handler.valid_ID(item['Country'], 'Countries',
                                 response=result)):
            return self.write(result)
        id = abs(item['Id'])
        if id !=0 and not handler.valid_ID(id, 'Tournaments', response=result):
            return self.write(result)
        if not handler.valid_ID(item.get('Owner', -1), 'Users', response=result,
                                msg='Invalid tournament owner'):
            return self.write(result)
        if item['Id'] >= 0 and item['End'] < item['Start']:
            result['message'] = 'End date must follow start date {}'.format(
                item['Start'])
            result['status'] = -7
            return self.write(result)
        result['message'] = 'Item {}'.format(
            'inserted' if id == 0 else 'deleted' if item['Id'] < 0 else
            'updated')
        if item['Id'] < 0:
            db.make_backup()
        try:
            with db.getCur() as cur:
                values = [item[f] for f in columns]
                if item['Id'] < 0:
                    sql = 'DELETE FROM Tournaments WHERE Id = ?'
                    args = (id, )
                elif item['Id'] == 0:
                    sql = 'INSERT INTO Tournaments ({}) VALUES ({})'.format(
                        ', '.join(columns),
                        ', '.join(['?' for v in values]))
                    args = values
                else:
                    sql = 'UPDATE Tournaments SET {} WHERE Id = ?'.format(
                        ', '.join('{} = ?'.format(f) for f in columns))
                    args = values + [id]
                log.debug('Executing "{}" on {}'.format(sql, args))
                cur.execute(sql, args)
                if item['Id'] == 0:
                    item['Id'] = cur.lastrowid
                    log.info('Last Tournament Row ID is now {}'.format(
                        item['Id']))
                item['Id'] = abs(item['Id']) # Put cleaned item record
                result['item'] = item # with correct Id in rsespone
        except sqlite3.DatabaseError as e:
            result['message'] = (
                'Exception in database change. SQL = {}. Args = {}. {}'.format(
                    sql, args, e))
            log.error(result['message'])
            result['status'] = -10

        return self.write(result)

def cleanTournamentItem(item, columns, current_user):
    global sql, args
    for field in ['Name', 'Location', 'Start', 'End', 'Logo', 'LinkURL']:
        if field in columns and item[field]:
            item[field] = item[field].strip()
    item['Owner'] = item.get('Owner', None) or current_user
    if 'Country' in columns and isinstance(item['Country'], str):
        with db.getCur() as cur:
            sql = 'SELECT Id FROM Countries WHERE Code = ?'
            args = (item['Country'].upper(), )
            cur.execute(sql, args)
            result = cur.fetchone()
            if result:
                item['Country'] = result[0]

class TournamentPlayerHandler(handler.BaseHandler):
    @tornado.web.authenticated
    def post(self):
        encoded_item = self.get_argument('item', None)
        item = json.loads(encoded_item)
        result = {'status': 0, 'message': ''}
        ID_fields = ('Id', 'Player', 'Tournament')
        update_fields = [f for f in db.table_field_names('Compete')
                         if f not in ID_fields]
        if not all(isinstance(item.get(field, None), int) 
                   for field in ID_fields):
            result['message'] = 'Invalid item in request'
            result['status'] = -1
            return self.write(result)
        id = abs(item['Id'])
        if id != 0 and not handler.valid_ID(
                id, 'Compete', response=result,
                msg='Invalid competitor in tournament.'):
            return self.write(result)
        if not handler.valid_ID(item['Player'], 'Players', response=result):
            return self.write(result)
        if not handler.valid_ID(item['Tournament'], 'Tournaments',
                                response=result):
            return self.write(result)
        id = abs(item['Id'])
        result['message'] = 'Competitor {}'.format(
            'inserted' if id == 0 else 'deleted' if item['Id'] < 0 else
            'updated')
        try:
            with db.getCur() as cur:
                sql = """SELECT Id FROM Compete WHERE Player = ? AND
                Tournament = ?"""
                args = (item['Player'], item['Tournament'])
                cur.execute(sql, args)
                matches = cur.fetchall()
                if len(matches) != 1 and id != 0:
                    result['message'] = 'No player {} in tournament {}'.format(
                        args)
                    result['status'] = -6
                elif len(matches) == 1 and id == 0:
                    result['message'] = (
                        'Player {} already in tournament {}'.format(args))
                    result['status'] = -6
                elif len(matches) == 1 and id != matches[0][0]:
                    result['message'] = (
                        'Internal error: Player {} already in tournament {} '
                        'with a different compete ID {}'.format(
                            args + matches[0]))
                    result['status'] = -6
                elif len(matches) > 1:
                    result['message'] = (
                        'Internal error: multiple compete records {}-{}'
                        .format(args))
                    result['status'] = -7
                    
                if result['status'] == 0:
                    if item['Id'] < 0:
                        sql = 'DELETE FROM Compete WHERE Id = ?'
                        args = (id, )
                    else:
                        fields = ['Player', 'Tournament'] + [
                            f for f in update_fields
                            if f in item and item[f] is not None]
                        if id > 0:
                            fields.append('Id')
                        values = [item[f] for f in fields]
                        sql = '''
                        INSERT OR REPLACE INTO Compete ({}) VALUES ({})
                        '''.format(', '.join(fields),
                                   ', '.join(['?' for v in values]))
                        args = values
                    log.debug('Executing "{}" on {}'.format(sql.strip(), args))
                    cur.execute(sql, args)
                    if id == 0:
                        item['Id'] = cur.lastrowid
                        log.debug('Last Compete row ID is now {}'.format(
                            item['Id']))
                    item['Id'] = abs(item['Id']) # Put cleaned item record
                    result['item'] = item # with correct Compete Id in respone
        except sqlite3.DatabaseError as e:
            result['message'] = (
                'Exception in database change. SQL = {}. Args = {}. {}'.format(
                    sql, args, e))
            log.error(result['message'])
            result['status'] = -10

        return self.write(result)

class EditTournamentHandler(handler.BaseHandler):
    tmt_fields = db.table_field_names('Tournaments')
    @tornado.web.authenticated
    def get(self, id=None):
        ctry_fields = db.table_field_names('Countries')
        tmt_fields = EditTournamentHandler.tmt_fields
        tmt_attr_fields = [f for f in tmt_fields if f != 'Id']
        with db.getCur() as cur:
            cur.execute("SELECT {} FROM Countries".format(
                ",".join(ctry_fields)))
            ctry_mapping = [(row[0], dict(zip(ctry_fields, row)))
                            for row in cur.fetchall()]
            countries = dict(ctry_mapping)
            def_country_id = ctry_mapping[0][1]["Id"]

            cur.execute("SELECT Id, Email FROM Users ORDER BY Email")
            users = dict(cur.fetchall())

            if id:
                cur.execute("SELECT {} FROM TOURNAMENTS WHERE Id = ?".format(
                    ",".join(tmt_fields)), (id,))
                results = cur.fetchall()
                if len(results) == 0:
                    return self.render(
                        "message.html",
                        message="Invalid tournament.",
                        next="Return to Tournament List",
                        next_url=settings.PROXYPREFIX)
                elif len(results) > 1:
                    return self.render(
                        "message.html",
                        message="ERROR multiple tournaments with ID {}."
                        "  Contact site administrator.".format(id),
                        next="Return to Tournament List",
                        next_url=settings.PROXYPREFIX)

                tournament = dict(zip(tmt_fields, results[0]))
                cur.execute("SELECT Name FROM Players"
                            " JOIN Compete on Compete.Player = Players.Id"
                            " WHERE Tournament = ?"
                            "  ORDER BY Name",
                            (id, ))
                tournament['Players'] = [row[0] for row in cur.fetchall()]
                cur.execute("SELECT COUNT(Scores.Id) FROM Scores"
                            "  JOIN Rounds ON Scores.Round = Rounds.Id"
                            "  WHERE Rounds.Tournament = ?",
                            (id,))
                tournament['ScoreCount'] = cur.fetchone()[0]
            else:
                tournament = dict(zip(tmt_fields, [''] * len(tmt_fields)))
                today = datetime.date.today()
                tomorrow = today + datetime.timedelta(days=1)
                tournament['Start'] = today.strftime(settings.DATEFORMAT)
                tournament['End'] = tomorrow.strftime(settings.DATEFORMAT)
                tournament['Country'] = def_country_id
                tournament['Owner'] = int(self.current_user)
                tournament['Players'] = []
                tournament['ScorePerPlayer'] = settings.DEFAULTSCOREPERPLAYER
                tournament['ScoreCount'] = 0
                sql = "INSERT INTO Tournaments ({}) VALUES ({})".format(
                    ', '.join(tmt_attr_fields),
                    ', '.join(['?'] * len(tmt_attr_fields)))
                try:
                    cur.execute(sql, [tournament[f] for f in tmt_attr_fields])
                    print('Created tournament', cur.lastrowid)
                    tournament['Id'] = cur.lastrowid
                except sqlite3.DatabaseError as e:
                    print('Error creating tournament ({}): {}'.format(sql, e))
                    return self.render(
                        'message.html',
                        message='Unable to create tournament: {}'.format(e))

            tournament['OwnerName'] = users[int(tournament['Owner'])]
            tournament['CountryCode'] = countries[tournament['Country']]['Code']
            tournament['CountryName'] = countries[tournament['Country']]['Name']
            tournament['Flag_Image'] = countries[tournament['Country']]['Flag_Image']

            if tournament['Owner'] == int(self.current_user) or self.get_is_admin():
                return self.render(
                    "edittournament.html",
                    tournament=tournament, users=users)
            else:
                return self.render(
                    "message.html",
                    message="ERROR only the tournament owner or administrators"
                    " may edit tournament attributes".format(id),
                    next="Return to Tournaments Home",
                    next_url=settings.PROXYPREFIX)

    @tornado.web.authenticated
    def post(self, id):
        tmt_fields = EditTournamentHandler.tmt_fields
        tmt_attr_fields = [f for f in tmt_fields if f != 'Id']
        state = {}
        for f in tmt_fields:
            state[f] = self.get_argument(f, None)
        if self.get_is_admin() or int(state['Owner']) == int(self.current_user):
            msg = 'Updated tournament {}'.format(state['Id'])
            if id == state['Id']:
                with db.getCur() as cur:
                    try:
                        for f in tmt_attr_fields:
                            sql = ("UPDATE Tournaments SET {} = ?"
                                   "  WHERE Id = ?").format(f)
                            cur.execute(sql, (state[f], id))
                        return self.write({'status': 'success', 'message': ''})
                    except sqlite3.DatabaseError as e:
                        print('Error updating tournament ({}): {}'.format(
                            sql, e))
                        return self.write({
                            'status': 'Error',
                            'message': 'Unable to update tournament {}: {}'
                            .format(id, e)})
            elif id and int(state['Id']) < 0:
                with db.getCur() as cur:
                    try:
                        sql = "DELETE FROM Tournaments WHERE Id = ?"
                        cur.execute(sql, (id,))
                        return self.write({
                            'status': 'load',
                            'message': 'Tournament Deleted',
                            'URL': settings.PROXYPREFIX})
                    except sqlite3.DatabaseError as e:
                        print('Error deleting tournament ({}): {}'.format(
                            sql, e))
                        return self.write({
                            'status': 'Error',
                            'message': 'Unable to update tournament {}: {}'
                            .format(id, e)})
            else:
                return self.write({'status': 'Error',
                                   'message': 'Inconsistent IDs {} and {}'.
                                   format(id, state['Id'])})
        else:
            self.write({'status': 'Error',
                        'message': 'You are not authorized to edit or create '
                        'that tournament.'})

class TournamentHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        tab = self.get_argument('tab', None)
        if tab:
            log.debug('Requested tab = {}'.format(tab))
        return self.render("tournament.html", tab=tab)

player_columns = ['Players.Id', 'Players.Name', 'Number', 'Countries.Code',
                  'Countries.Id', 'Flag_Image',
                  'Association', 'Pool', 'Type', 'Wheel']
player_fields = ["id", "name", "number", "country", 
                 "countryid", "flag_image",
                 "association", "pool", "type", "wheel"]

def getPlayers(tournamentid):
    global player_fields
    with db.getCur() as cur:
        sql = """
        SELECT {} FROM Players
          LEFT OUTER JOIN Countries
            ON Countries.Id = Players.Country
          LEFT JOIN Compete ON Compete.Player = Players.Id
          INNER JOIN Tournaments
            ON Tournaments.Id = Compete.Tournament
        WHERE Tournaments.Id = ?
        ORDER BY Players.Name ASC
        """.format(','.join(player_columns)).strip()
        args = (tournamentid,)
        cur.execute(sql, args)
        players = [dict(zip(player_fields, row)) for row in cur.fetchall()]
        for player in players:
            player['type'] = db.playertypes[int(player['type'] or 0)]
    return players

class ShowPlayersHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        players = getPlayers(self.tournamentid)
        editable = self.get_is_admin() or (
            self.current_user and self.current_user == str(self.owner))
        return self.render("players.html", editable = editable,
                           players = players)

class DeletePlayerHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        player = self.get_argument("player", None)
        if player is None:
            return self.write({'status':"error",
                               'message':"Please provide a player"})
        try:
            with db.getCur() as cur:
                if player == "all":
                    cur.execute("DELETE FROM Compete WHERE Tournament = ?",
                                (self.tournamentid,))
                else:
                    cur.execute("DELETE FROM Compete WHERE Player = ?"
                                " AND Tournament = ?",
                                (player, self.tournamentid))
                return self.write({'status':"success"})
        except:
            return self.write({'status':"error",
                 'message':"Couldn't delete player from tournament"})

class PlayersHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    def get(self):
        data = {'players': getPlayers(self.tournamentid),
                'editable': self.get_is_admin() or (
                    self.current_user and
                    self.current_user == str(self.owner))
                }
        return self.write(data)
    
    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        global player_fields
        player = self.get_argument("player", None)
        if player is None or not (player.isdigit() or player == '-1'):
            return self.write({'status':"error", 'message':"Please provide a player"})
        info = self.get_argument("info", None)
        if info is None:
            return self.write({'status':"error", 'message':"Please provide an info object"})
        info = json.loads(info)
        try:
            fields = []
            with db.getCur() as cur:
                for colname, val in info.items():
                    col = colname.lower()
                    if not (col in player_fields and
                            (db.valid[col if col in db.valid else 'all'].match(
                                val))):
                        fields.append(col)
                    if player == '-1':
                        cur.execute("INSERT INTO Players (Name, Country, Tournament) VALUES"
                                    " ('\u202Fnewplayer',"
                                    "  (select Id from Countries limit 1), ?)", (self.tournamentid,))
                    else:
                        if colname == "Type" and val == str(
                                db.playertypecode['Substitute']):
                            cur.execute(
                                "UPDATE Players SET Country ="
                                "   (SELECT Id FROM Countries WHERE"
                                "      Name = 'Substitute' OR 'Code' = 'SUB'"
                                "             OR IOC_Code = 'SUB')"
                                " WHERE Id = ? AND Tournament = ?",
                                (player, self.tournamentid))
                        cur.execute("UPDATE Players SET {0} = ? WHERE Id = ? AND Tournament = ?"
                                    .format(colname),
                                    (val, player, self.tournamentid))
            if len(fields) > 0:
                return self.write(
                    {'status':"error",
                     'message':
                     "Invalid column(s) or value provided: {0}".format(
                         ", ".join(fields))})
            return self.write({'status':"success"})
        except:
            return self.write({'status':"error",
                 'message':"Invalid info provided"})

SS_Player_Columns = [f.capitalize() for f in player_fields
                     if f not in ('id', 'countryid', 'flag_image')]
excel_mime_type = '''
application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'''.strip()

class DownloadTournamentSheetHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        players = getPlayers(self.tournamentid)
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = 'Players'
        header_row = 3
        first_column = 1
        top_center_align = openpyxl.styles.Alignment(
            horizontal='center', vertical='top', wrap_text=True)
        sheet.merge_cells(
            start_row=header_row - 2, end_row=header_row - 2,
            start_column=first_column, 
            end_column=first_column + len(SS_Player_Columns) - 1)
        title_cell = sheet.cell(
            row=header_row - 2, column=first_column,
            value='{} Players'.format(self.tournamentname))
        title_cell.alignment = top_center_align
        row = header_row
        for i, column in enumerate(SS_Player_Columns):
            sheet.cell(row=row, column=first_column + i,
                       value=column)
        for player in players:
            row += 1
            for i, f in enumerate(SS_Player_Columns):
                sheet.cell(row=row, column=first_column + i,
                           value=player[f.lower()])
        with tempfile.NamedTemporaryFile(
                suffix='.xlsx',
                prefix='{}_'.format(self.tournamentname)) as outf:
            book.save(outf.name)
            outf.seek(0)
            self.write(outf.read())
            self.set_header('Content-type', excel_mime_type)
            self.set_header(
                'Content-Disposition',
                'attachment; filename="{}"'.format(os.path.basename(outf.name)))
        return

class UploadPlayersHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        if 'file' not in self.request.files or len(self.request.files['file']) == 0:
            return self.write({'status':"error", 'message':"Please provide a players file"})
        players = self.request.files['file'][0]['body']
        colnames = [c.lower() for c in SS_Player_Columns]
        try:
            with db.getCur() as cur:
                reader = csv.reader(players.decode('utf-8').splitlines())
                good = 0
                bad = 0
                first = True
                unrecognized = []
                for row in reader:
                    hasheader = first and sum(
                        1 if v.lower() in colnames else 0 for v in row) >= len(row) - 1
                    first = False
                    if hasheader:
                        unrecognized = [v for v in row
                                        if v.lower() not in colnames]
                        colnames = [v.lower() for v in row]
                        colnames += [c.lower() for c in SS_Player_Columns
                                     if c.lower() not in colnames]
                        continue
                    if len(row) < 3:
                        bad += 1;
                        continue
                    filled = 0
                    for i in range(len(row)):
                        if row[i] == '':
                            row[i] = None
                        else:
                            filled += 1
                    if filled < 3:
                        bad += 1
                        continue
                    rowdict = dict(zip(
                        colnames, list(row) + [None] * len(colnames)))
                    if not (rowdict['name'] or
                            (rowdict['number'] and
                             not rowdict['number'].isdigit()) or
                            (rowdict['type'] and
                             not rowdict['type'] in db.playertypes)):
                        bad += 1
                        continue
                    rowdict['type'] = db.playertypecode[rowdict['type']] if (
                        rowdict['type'] in db.playertypecode) else 0
                    cur.execute(
                        "INSERT INTO Players(Name, Number, Country, "
                        "                    Association, Pool, Type, Wheel, Tournament)"
                        " VALUES(?, ?,"
                        "   (SELECT Id FROM Countries"
                        "      WHERE Name = ? OR Code = ? OR IOC_Code = ?),"
                        "   ?, ?, ?, ?, ?);",
                        (rowdict['name'], rowdict['number'], rowdict['country'],
                         rowdict['country'], rowdict['country'],
                         rowdict['association'], rowdict['pool'],
                         rowdict['type'], rowdict['wheel'], self.tournamentid))
                    good += 1
            message = ("{} player record(s) loaded, "
                       "{} player record(s) skipped").format(good,bad)
            if unrecognized:
                message += ", skipped column header(s): {}".format(
                    ', '.join(unrecognized))
            return self.write({'status':"success", 'message': message})
        except Exception as e:
            return self.write({'status':"error",
                    'message':"Invalid players file provided: " + str(e)})


class AddRoundHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        with db.getCur() as cur:
            print(self.tournamentid)
            cur.execute(
                "INSERT INTO Rounds(Number, Name, Tournament) "
                "VALUES((SELECT COUNT(*) + 1 FROM Rounds WHERE Tournament = ?),"
                "       'Round ' || (SELECT COUNT(*) + 1 FROM Rounds "
                "                    WHERE Tournament = ?),"
                "       ?)", (self.tournamentid,) * 3)
            return self.write({'status':"success"})

class DeleteRoundHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        round = self.get_argument("round", None)
        if round is None:
            return self.write({'status':"error", 'message':"Please provide a round"})
        with db.getCur() as cur:
            cur.execute("DELETE FROM Rounds WHERE Id = ? AND Tournament = ?", (round, self.tournamentid))
            return self.write({'status':"success"})

def getSettings(self, tournamentid):
    with db.getCur() as cur:
        cur.execute("""SELECT Id, Number, Name, COALESCE(Ordering, 0),
                        COALESCE(Algorithm, 0), Seed,
                        Cut, SoftCut, CutSize, CutMobility,
                        CombineLastCut, CutCount,
                        Duplicates, Diversity, UsePools, Winds, Games
                    FROM Rounds WHERE Tournament = ?
                    ORDER BY Number""", (tournamentid,))

        cols = ["id", "number", "name", "ordering", "algorithm", "seed",
                "cut", "softcut", "cutsize", "cutmobility",
                "combinelastcut", "cutcount",
                "duplicates", "diversity", "usepools", "winds", "games"]
        rounds = []

        for row in cur.fetchall():
            roundDict = dict(zip(cols, row))

            roundDict["orderingname"] = seating.ORDERINGS[roundDict["ordering"]][0]
            roundDict["algname"] = seating.ALGORITHMS[roundDict["algorithm"]].name
            roundDict["seed"] = roundDict["seed"] or ""

            rounds += [roundDict]

        cur.execute('SELECT COALESCE(ScorePerPlayer, {}) FROM Tournaments'
                    ' WHERE Id = ?'.format(settings.DEFAULTSCOREPERPLAYER),
                    (tournamentid,))
        result = cur.fetchone()
        if result is None:
            return None
        scorePerPlayer = result[0]
        return {'rounds':rounds,
                'scoreperplayer': scorePerPlayer,
                'unusedscoreincrement': settings.UNUSEDSCOREINCREMENT,
                'cutsize':settings.DEFAULTCUTSIZE}
    return None

class SettingsHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    def get(self):
        return self.write(getSettings(self, self.tournamentid))
    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        round = self.get_argument("round", None)
        if round is None:
            return self.write({'status':"error", 'message':"Please provide a round"})
        settings = self.get_argument("settings", None)
        if settings is None:
            return self.write({'status':"error", 'message':"Please provide a settings object"})
        settings = json.loads(settings)
        print(settings)
        with db.getCur() as cur:
            for colname, val in settings.items():
                cur.execute("UPDATE Rounds SET {0} = ? WHERE Id = ? AND Tournament = ?".format(colname),
                        (val, round, self.tournamentid)) # TODO: fix SQL injection
            return self.write({'status':"success"})


class ShowSettingsHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        roundsettings = getSettings(self, self.tournamentid)
        return self.render("roundsettings.html", rounds=roundsettings['rounds'], cutsize=roundsettings['cutsize'])

def getTourney(self, tournamentid):
    stage_fields = db.table_field_names('Stages')
    with db.getCur() as cur:
        cur.execute("""SELECT {} FROM Stages WHERE Tournament = ?
                       ORDER BY SortOrder, Id""".format(",".join(stage_fields)),
                    (tournamentid,))
        stages = dict((row[0], dict(zip(stage_fields, row)))
                      for row in cur.fetchall())
        roots = [id for id in stages if not stages[id]['PreviousStage']]
        if len(roots) > 1:
            raise Exception("Tournament {} has multiple initial stages".format(
                tournamentid))
        elif len(roots) == 0 and len(stages) > 0:
            raise Exception("Tournament {} has no initial stage but a cycle "
                            "of others".format(tournamentid))
        elif len(roots) == 0:
            stage = {'Tournament': tournamentid,
                     'Name': 'Round Robin',
                     'SortOrder': 0,
                     'PreviousStage': None,
                     'Ranks': None,
                     'Cumulative': 0,}
            cur.execute("""INSERT INTO Stages ({}) VALUES ({})""".format(
                ','.join(f for f in stage if stage[f] is not None),
                ','.join(repr(stage[f]) for f in stage
                         if stage[f] is not None)))
            stage['Id'] = cur.lastrowid
            stages = {cur.lastrowid: stage}
            roots = [cur.lastrowid]

        # List stages starting with root and then successor children, in order
        stagelist = [stages[roots[0]]]
        todo=stages.copy()
        del todo[roots[0]]
        while len(todo) > 0:
            children = [id for id in stages
                        if stages[id]['PreviousStage'] in roots]
            for child in sorted(children, key=lambda s: s['SortOrder']):
                stages[child]['previousName'] = stages[
                    stages[child]['PreviousStage']]['Name']
                stagelist.append(stages[child])
                del todo[child]
                roots.append(child)

        cols = ["id", "number", "name", "ordering", "algorithm", "seed",
                "cut", "softcut", "cutsize", "cutmobility", "combinelastcut",
                "duplicates", "diversity", "usepools", "winds", "games"]
        for stage in stagelist[:1]:
            cur.execute("""SELECT Id, Number, Name, COALESCE(Ordering, 0),
                             COALESCE(Algorithm, 0), Seed,
                             Cut, SoftCut, CutSize, CutMobility, CombineLastCut,
                             Duplicates, Diversity, UsePools, Winds, Games
                           FROM Rounds WHERE Tournament = ?
                           ORDER BY Number""", (tournamentid,))

            stage['rounds'] = []

            for row in cur.fetchall():
                roundDict = dict(zip(cols, row))
                roundDict["orderingname"] = seating.ORDERINGS[
                    roundDict["ordering"]][0]
                roundDict["algname"] = seating.ALGORITHMS[
                    roundDict["algorithm"]].name
                roundDict["seed"] = roundDict["seed"] or ""
                stage['rounds'] += [roundDict]

        cur.execute('SELECT COALESCE(ScorePerPlayer, {}) FROM Tournaments'
                    ' WHERE Id = ?'.format(settings.DEFAULTSCOREPERPLAYER),
                    (tournamentid,))
        result = cur.fetchone()
        if result is None:
            return None
        scorePerPlayer = result[0]
        return {'stages': stagelist, 'scoreperplayer': scorePerPlayer,
                'unusedscoreincrement': settings.UNUSEDSCOREINCREMENT,
                'cutsize':settings.DEFAULTCUTSIZE}

class TourneySettingsAjaxHandler(handler.BaseHandler):
    @handler.tournament_handler_ajax
    def get(self):
        return self.write(getTourney(self, self.tournamentid))

    @handler.tournament_handler_ajax
    @handler.is_owner_ajax
    def post(self):
        stage = self.get_argument("stage", None)
        if stage is None:
            return self.write({'status':"error",
                               'message':"Please provide a stage"})
        settings = self.get_argument("settings", None)
        if settings is None:
            return self.write({'status':"error",
                               'message':"Please provide a settings object"})
        settings = json.loads(settings)
        # print(settings)
        with db.getCur() as cur:
            for colname, val in settings.items():
                cur.execute("UPDATE Stages SET {} = ?"
                            " WHERE Id = ? AND Tournament = ?".format(colname),
                        (val, stage, self.tournamentid)) # TODO: fix SQL injection
            return self.write({'status':"success"})

class TourneySettingsHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        settings = getTourney(self, self.tournamentid)
        return self.render("tourney.html", **settings)

countryColumns = ('Id', 'Name', 'Code', 'IOC_Code', 'IOC_Name', 'Flag_Image')
def getCountries():
    countries = []
    with db.getCur() as cur:
        cur.execute("SELECT {cols} FROM Countries".format(
            cols=",".join(countryColumns)))
        countries = [dict(zip(countryColumns, row)) for row in cur.fetchall()]
        countries.sort(key=lambda country: country['Code'])
        for j in range(len(countries)):
            if countries[j]['Code'] == 'SUB':
                sub = countries[j]
                countries[j:j+1] = []
                countries.append(sub)
                break
    return countries

class CountriesHandler(handler.BaseHandler):
    def get(self):
        return self.write(json.dumps(getCountries()))

class UpdateCountriesHandler(handler.BaseHandler):
    @handler.is_admin
    def get(self):
        countriesDB = dict((country['Code'], country) 
                           for country in getCountries())
        countriesFile = dict((country['Code'], country) 
                             for country in db.getCountriesFromFile())
        adds, mods = [], []
        fields = [f for f in countryColumns if f not in ('Id', 'Code')]
        updateSQL = "UPDATE Countries SET {} WHERE Id = ?".format(
            ', '.join('{} = ?'.format(f) for f in fields))
        insertSQL = "INSERT INTO Countries ({}) VALUES ({})".format(
            ', '.join(fields + ['Code']),
            ', '.join('?' for f in fields + ['Code']))
        with db.getCur() as cur:
            for code in countriesFile:
                fileCountry = countriesFile[code]
                dbCountry = countriesDB.get(code, None)
                if dbCountry and util.dict_differences( 
                        fileCountry, dbCountry, fields):
                    args = [fileCountry[f] for f in fields] + [dbCountry['Id']]
                    cur.execute(updateSQL, args)
                    mods.append(code)
                elif dbCountry is None:
                    args = [fileCountry[f] for f in fields + ['Code']]
                    cur.execute(insertSQL, args)
                    adds.append(code)
        return self.render(
            'message.html',
            message=('<p>Among {} countries in the database:</p>'
                     '<p>Countries updated ({}):</p> <p>{}</p>'
                     '<p>Countries inserted ({}):</p> <p>{}</p>'.format(
                         len(countriesDB), len(mods), mods, len(adds), adds)),
            next='Return to Tournaments Home',
            next_url=settings.PROXYPREFIX)

class AssociationsHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        with db.getCur() as cur:
            cur.execute("SELECT DISTINCT Association FROM Players"
                        " LEFT JOIN Compete ON Compete.Player = Players.Id"
                        " WHERE Association IS NOT null"
                        "       AND length(Association) > 0"
                        "       AND Tournament = ?", (self.tournamentid,))
            return self.write(json.dumps([row[0] for row in cur.fetchall()]))
