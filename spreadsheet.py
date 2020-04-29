#!/usr/bin/env python3

import re
import logging
import sqlite3
import os.path
from collections import *
import math
import io

import tornado.web
import openpyxl
import tempfile
import handler
import db
import tournament
import seating
import leaderboard
import util

log = logging.getLogger('WebServer')

Player_Columns = [f.capitalize() for f in tournament.player_fields
                     if f not in ('id', 'countryid', 'flag_image')] + [
                             'Tournament']
excel_mime_type = '''
application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'''.strip()

# OpenPyXL formatting definitions
top_center_align = openpyxl.styles.Alignment(
   horizontal='center', vertical='top', wrap_text=True)
default_font = openpyxl.styles.Font()
title_font = openpyxl.styles.Font(name=default_font.name, size=14, bold=True)
column_header_font = openpyxl.styles.Font(
    name=default_font.name, size=12, bold=True)
no_border = openpyxl.styles.Border()
thin_outline = openpyxl.styles.Border(
   outline=openpyxl.styles.Side(border_style="thin")
)
paleYellowFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="FFFFCC", fill_type='solid')
paleGreenFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="DDFFDD", fill_type='solid')
paleBlueFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="DDDDFF", fill_type='solid')
blackFill = openpyxl.styles.fills.PatternFill(
    patternType='solid', fgColor="000000", fill_type='solid')

def merge_cells(sheet, row, column, height=1, width=1, 
                font=title_font, align=top_center_align, border=no_border,
                value=None, fill=None):
   if height > 1 or width > 1:
      sheet.merge_cells(
         start_row=row, start_column=column,
         end_row=row + height -1, end_column=column + width - 1)
   top_left = sheet.cell(row, column)
   top_left.alignment = align
   top_left.font = font
   top_left.border = border
   if value:
       top_left.value = value
   if fill:
       top_left.fill = fill
   return top_left

def inMergedCell(cell, sheet):
    "Test if cell is within a merged cell of the sheet"
    for mergedCellRange in sheet.merged_cells.ranges:
        if (mergedCellRange.min_col <= cell.column and
            cell.column <= mergedCellRange.max_col and
            mergedCellRange.min_row <= cell.row and
            cell.row <= mergedCellRange.max_row):
            return True
    return False

def resizeColumn(sheet, column, min_row=None, max_row=None, 
                 character_width=1.01, min_width=0, exclude_merged=True):
    base_font_size = 10
    width = min_width * base_font_size
    for tup in sheet.iter_rows(
            min_col=column, max_col=column, min_row=min_row, max_row=max_row):
        cell = tup[0]
        if cell.value and not (exclude_merged and inMergedCell(cell, sheet)):
            font_size = cell.font.sz if cell.font and cell.font.sz else 10
            width = max(width, len(str(cell.value)) * font_size)
    letter = openpyxl.utils.cell.get_column_letter(column)
    sheet.column_dimensions[letter].width = math.ceil(
        width * character_width / base_font_size)
        
def makePlayersSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet()
    sheet.title = 'Players'
    players = tournament.getPlayers(tournamentID)
    header_row = 3
    first_column = 1
    row = header_row
    columns = Player_Columns
    merge_cells(sheet, header_row - 2, first_column, 1, len(columns),
                font=title_font, border=thin_outline,
                value='{} Players'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    for i, column in enumerate(columns):
        cell = sheet.cell(row, first_column + i, value = column)
        cell.font = column_header_font
        cell.alignment = top_center_align
    for player in players:
        row += 1
        for i, f in enumerate(columns):
            cell = sheet.cell(
                row=row, column=first_column + i,
                value=tournamentName if f == 'Tournament'
                else player[f.lower()])
    for col in range(first_column, first_column + len(columns)):
        resizeColumn(sheet, col, min_row=header_row)
    return sheet

def makeSettingsSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'Settings'
    tmt_fields = [f for f in db.table_field_names('Tournaments')
                  if f not in ('Id', 'Name', 'Country', 'Owner')]
    rounds_fields = [f for f in db.table_field_names('Rounds') 
                     if f not in ('Id', 'Tournament')]
    query_fields = ['Countries.Code', 'Email'] + [
        'Tournaments.{}'.format(f) for f in tmt_fields] + [
        'Rounds.{}'.format(f) for f in rounds_fields]
    round_display_fields = [
        'Name', 'Number', 'Ordering', 'Algorithm', 'CutSize', 'Games']
    sql = """
    SELECT {} FROM Tournaments
      LEFT OUTER JOIN Countries ON Countries.Id = Tournaments.Country
      LEFT OUTER JOIN Users ON Users.Id = Tournaments.Owner
      LEFT OUTER JOIN Rounds ON Rounds.Tournament = Tournaments.Id
    WHERE Tournaments.Id = ?
    """.strip().format(','.join(query_fields))
    args = (tournamentID,)
    with db.getCur() as cur:
        cur.execute(sql, args)
        rounds = [dict(zip(map(db.fieldname, query_fields), row))
                  for row in cur.fetchall()]
    header_row = 3
    first_column = 1
    row = header_row
    merge_cells(sheet, header_row - 2, first_column, 
                1, max(2, len(round_display_fields)),
                font=title_font, border=thin_outline,
                value='{} Settings'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    top_left = first_column + (len(round_display_fields) - 2) // 2
    for field in tmt_fields + ['Code']:
        if field not in ('Name', ):
            namecell = sheet.cell(
                row, top_left, value='Country ' + field if field == 'Code' else
                'Owner ' + field if field == 'Email' else field)
            valuecell = sheet.cell(row, top_left + 1, value=rounds[0][field])
            row += 1
    row += 2
    for i, field in enumerate(round_display_fields):
        cell = sheet.cell(
            row, first_column + i,
            value = 'Seating Algorithm' if field == 'Algortihm' else
            'Cut Size' if field == 'CutSize' else
            'Round Name' if field == 'Name' else
            'Round Number' if field == 'Number' else
            field)
        cell.font = column_header_font
        cell.alignment = top_center_align
    if len(rounds) == 1 and not rounds[0]['Name']:
        row += 1
        merge_cells(sheet, row, first_column, 1,
                    len(round_display_fields), font=default_font,
                    value='No rounds defined')
    else:
        for round in rounds:
            row += 1
            for i, field in enumerate(round_display_fields):
                cell = sheet.cell(
                    row, first_column + i,
                    value = seating.ALGORITHMS[round[field] or 0].name
                    if field == 'Algorithm' else
                    seating.ORDERINGS[round[field] or 0][0]
                    if field == 'Orderings' else
                    round[field])
    for col in range(first_column, 
                     first_column + max(2, len(round_display_fields))):
        resizeColumn(sheet, col, min_row = header_row)
    return sheet

def makeScoresSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'All Scores'
    player_fields = ['Name', 'Country', 'Status']
    round_display_fields = ['Rank', 'Points', 'Penalty', 'Total']
    scoreboard, rounds = leaderboard.getTournamentScores(tournamentID)
    header_row = 3
    first_column = 1
    total_columns = len(player_fields) + len(round_display_fields) * len(rounds)
    row = header_row
    roundColorFills = [paleGreenFill, paleBlueFill]
    merge_cells(sheet, header_row - 2, first_column, 
                1, min(total_columns, 20),
                font=title_font, border=thin_outline,
                value='{} Scores'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    for i, field in enumerate(player_fields):
        cell = sheet.cell(row + 1, first_column + i, value=field)
        cell.font = column_header_font
        cell.alignment = top_center_align
    col = first_column + len(player_fields)
    color = 0
    for roundID, roundName in rounds:
        round_cell = merge_cells(
            sheet, row, col, 1, len(round_display_fields), value=roundName,
            fill=roundColorFills[color])
        for j, rfield in enumerate(round_display_fields):
            cell = sheet.cell(row + 1, col + j, value=rfield)
            cell.font = column_header_font
            cell.alignment = top_center_align
            cell.fill = roundColorFills[color]
        col += len(round_display_fields)
        color = 1 - color
    row += 1
    for player in scoreboard:
        row += 1
        for i, field in enumerate(player_fields):
            cell = sheet.cell(
                row, first_column + i, 
                value=player['type' if field == 'Status' else field.lower()])
        col = first_column + len(player_fields)
        color = 0
        for roundID, roundName in rounds:
            for j, rfield in enumerate(round_display_fields):
                cell = sheet.cell(row, col + j)
                cell.fill = roundColorFills[color]
                if roundID in player['scores']:
                    cell.value=player['scores'][roundID][
                        'score' if rfield == 'Points' else rfield.lower()]
            col += len(round_display_fields)
            color = 1 - color
        
    for col in range(first_column, first_column + total_columns):
        resizeColumn(sheet, col, min_row = header_row)
    return sheet

def makeStandingsSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'Standings'
    players, allTied = leaderboard.leaderData(tournamentID)
    fields = ([] if allTied else ['Place']) + [
        'Name', 'Country', 'Status', 'Games', 'Points', 'Penalty', 'Total']
    header_row = 3
    first_column = 1
    row = header_row
    merge_cells(sheet, header_row - 2, first_column, 1, len(fields),
                font=title_font, border=thin_outline,
                value='{} Standings'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    for i, field in enumerate(fields):
        cell = sheet.cell(
            row, first_column + i,
            value = 'Raw Points' if field == 'Points' else field)
        cell.font = column_header_font
        cell.alignment = top_center_align
    last_cut = None
    for player in players:
        row += 1
        if player['cutName'] and player['cutName'] != last_cut:
            separator = merge_cells(sheet, row, first_column, 1, len(fields),
                                    border=thin_outline, fill = blackFill)
            sheet.row_dimensions[row].height = title_font.size // 3
            row += 1
            cutname = merge_cells(sheet, row, first_column + 1,
                                  1, len(fields) - 2,
                                  border=thin_outline, font=default_font,
                                  value = 'CUT ' + player['cutName'],
                                  fill = paleYellowFill)
            for col in (first_column, first_column + len(fields) - 1):
                sheet.cell(row = row, column = col).fill = paleYellowFill
            row += 1
            last_cut = player['cutName']
        for i, field in enumerate(fields):
            cell = sheet.cell(
                row, first_column + i,
                value=player['gamesPlayed' if field == 'Games' else
                             'type' if field == 'Status' else
                             field.lower()])
    for col in range(first_column, first_column + len(fields)):
        resizeColumn(sheet, col, min_row = header_row)
    return sheet
        
def makeSeatingAndScoresSheet(book, tournamentID, tournamentName, sheet=None):
    if sheet is None:
        sheet = book.create_sheet() 
    sheet.title = 'Seating & Scores'
    rounds = seating.getSeating(tournamentID)
    fields = (('Wind', ) if rounds and rounds[0]['winds'] else tuple()) + (
        'Name', 'Assoc.', 'Country', 'RawScore', 'Rank', 'Score', 'Penalty')
    header_row = 3
    first_column = 1
    row = header_row
    merge_cells(sheet, header_row - 2, first_column, 
                1, (len(fields) + 1) * max(1, len(rounds)),
                font=title_font, border=thin_outline,
                value='{} Seating & Scores'.format(tournamentName))
    sheet.row_dimensions[header_row - 2].height = title_font.size * 3 // 2
    if len(rounds) == 0:
        merge_cells(sheet, row, first_column, 1, len(fields),
                    font=column_header_font,
                    value='No seating or scores found')
    roundColorFills = [paleGreenFill, paleBlueFill]
    color = 0
    rounds.sort(key=lambda r: r['round'])
    for r, round in enumerate(rounds):
        col = first_column + r * (len(fields) + 1)
        row = header_row
        merge_cells(sheet, row, col, 1, len(fields),
                    font=column_header_font,
                    value=round['name'], fill=roundColorFills[color])
        row += 1
        if len(round['tables']) == 0:
            merge_cells(sheet, row, col, 1, len(fields),
                        value='No seating or scores found',
                        fill=roundColorFills[color])
            row += 1
        last_cut = None
        for table in sorted(round['tables'], key=lambda t: t['table']):
            if table['cutName'] and table['cutName'] != last_cut:
                separator = merge_cells(sheet, row, col, 1, len(fields),
                                        border=thin_outline, fill = blackFill)
                # sheet.row_dimensions[row].height = title_font.size // 3
                row += 1
                cutname = merge_cells(sheet, row, col, 1, len(fields),
                                      border=thin_outline, font=default_font,
                                      value = 'CUT ' + table['cutName'],
                                      fill = paleYellowFill)
                row += 1
                last_cut = table['cutName']
            for c in range(col, col + len(fields)):
                sheet.cell(row, c).fill = roundColorFills[color]
            row += 1
            table_name = merge_cells(
                sheet, row, col, 1, len(fields),
                border=thin_outline, font=column_header_font,
                value = 'Table {}'.format(table['table']),
                fill=roundColorFills[color])
            row += 1
            for i, field in enumerate(fields):
                cell = sheet.cell(
                    row, col + i,
                    value='Raw Score' if field == 'RawScore' else field)
                cell.alignment = top_center_align
                cell.fill = roundColorFills[color]
            row += 1
            for wind_and_player in table['players']:
                for i, field in enumerate(fields):
                    cell = sheet.cell(
                        row, col + i,
                        value=wind_and_player['wind'] if field == 'Wind' else
                        wind_and_player['player'][
                            'association' if field == 'Assoc.' else
                            field.lower()])
                    cell.fill = roundColorFills[color]
                row += 1
            if table['unusedPoints']['rawscore']:
                extra = 1 if 'Wind' in fields else 0
                merge_cells(
                    sheet, row, col, 1, 3 + extra, font=default_font,
                    value='Unused Points', fill = roundColorFills[color])
                merge_cells(
                    sheet, row, col + 3 + extra, 1, 4, font=default_font,
                    value=table['unusedPoints']['rawscore'],
                    fill = roundColorFills[color])
                row += 1
        color = 1 - color
    # for col in range(first_column + len(fields), len(rounds) * len(fields),
    #                  len(fields) + 1):
    #     letter = openpyxl.utils.cell.get_column_letter(col)
    #     sheet.column_dimensions[letter].width = 2
    for col in range(first_column, 
                     first_column + max(1, len(rounds)) * (len(fields) + 1)):
       resizeColumn(sheet, col, min_row=header_row, min_width=2)
    return sheet
        
class DownloadTournamentSheetHandler(handler.BaseHandler):
    @handler.tournament_handler
    def get(self):
        book = openpyxl.Workbook()
        standingsSheet = makeStandingsSheet(
            book, self.tournamentid, self.tournamentname, book.active)
        seatingAndScoresSheet = makeSeatingAndScoresSheet(
            book, self.tournamentid, self.tournamentname)
        scoresSheet = makeScoresSheet(
            book, self.tournamentid, self.tournamentname)
        playersSheet = makePlayersSheet(
            book, self.tournamentid, self.tournamentname)
        settingsSheet = makeSettingsSheet(
            book, self.tournamentid, self.tournamentname)
        with tempfile.NamedTemporaryFile(
                suffix='.xlsx',
                prefix='{}_'.format(util.makeFilename(self.tournamentname))
        ) as outf:
            book.save(outf.name)
            outf.seek(0)
            self.write(outf.read())
            self.set_header('Content-type', excel_mime_type)
            self.set_header(
                'Content-Disposition',
                'attachment; filename="{}"'.format(os.path.basename(outf.name)))
            log.debug('Temporary file: {}'.format(outf.name))
        return

class UploadPlayersHandler(handler.BaseHandler):
    @handler.is_authenticated_ajax
    def post(self):
        response = {'status': 0, 'message': 'Player data received'}
        return self.write(response)

class FindPlayersInSpreadsheetHandler(handler.BaseHandler):
    @handler.is_authenticated_ajax
    def post(self):
        if not ('file' in self.request.files and
                len(self.request.files['file']) > 0):
            return self.write({'status': -1,
                               'message':"Please provide a players file"})
        response = {'status': 0, 'message': 'Valid spreadsheet received'}
        content = self.request.files['file'][0]['body']
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, keep_vba=False,
                data_only=True, keep_links=False)
            playerLists = []
            for sheet in workbook:
                playerLists.extend(find_players_in_sheet(sheet))
            playerLists.sort(key=lambda pl: len(pl['players']),
                             reverse=True)
            response['message'] = 'Found {} player lists'.format(
                len(playerLists))
            response['playerLists'] = playerLists
        except Exception as e:
            response['status'] = -1
            response['message'] = 'Error processing spreadsheet. {}'.format(e)

        return self.write(response)

def find_players_in_sheet(sheet):
    result = []
    if sheet.title == 'Players':
        result.append({'sheet': sheet.title, 'players': [], 'top': 'A1'})
    return result

def process():
        try:
            with db.getCur() as cur:
                reader = csv.reader(players.decode('utf-8').splitlines())
                good = 0
                bad = 0
                first = True
                unrecognized = []
                for row in reader:
                    hasheader = first and sum(
                        1 if v.lower() in colnames else 0 for v in row) >= len(row) - 1
                    first = False
                    if hasheader:
                        unrecognized = [v for v in row
                                        if v.lower() not in colnames]
                        colnames = [v.lower() for v in row]
                        colnames += [c.lower() for c in Player_Columns
                                     if c.lower() not in colnames]
                        continue
                    if len(row) < 3:
                        bad += 1;
                        continue
                    filled = 0
                    for i in range(len(row)):
                        if row[i] == '':
                            row[i] = None
                        else:
                            filled += 1
                    if filled < 3:
                        bad += 1
                        continue
                    rowdict = dict(zip(
                        colnames, list(row) + [None] * len(colnames)))
                    if not (rowdict['name'] or
                            (rowdict['number'] and
                             not rowdict['number'].isdigit()) or
                            (rowdict['type'] and
                             not rowdict['type'] in db.playertypes)):
                        bad += 1
                        continue
                    rowdict['type'] = db.playertypecode[rowdict['type']] if (
                        rowdict['type'] in db.playertypecode) else 0
                    cur.execute(
                        "INSERT INTO Players(Name, Number, Country, "
                        "                    Association, Pool, Type, Wheel, Tournament)"
                        " VALUES(?, ?,"
                        "   (SELECT Id FROM Countries"
                        "      WHERE Name = ? OR Code = ? OR IOC_Code = ?),"
                        "   ?, ?, ?, ?, ?);",
                        (rowdict['name'], rowdict['number'], rowdict['country'],
                         rowdict['country'], rowdict['country'],
                         rowdict['association'], rowdict['pool'],
                         rowdict['type'], rowdict['wheel'], self.tournamentid))
                    good += 1
            message = ("{} player record(s) loaded, "
                       "{} player record(s) skipped").format(good,bad)
            if unrecognized:
                message += ", skipped column header(s): {}".format(
                    ', '.join(unrecognized))
            return self.write({'status':"success", 'message': message})
        except Exception as e:
            return self.write({'status':"error",
                    'message':"Invalid players file provided: " + str(e)})
